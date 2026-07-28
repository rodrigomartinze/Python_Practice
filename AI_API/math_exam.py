import os
import json
import ast
import time
from groq import Groq
from dotenv import load_dotenv
import pandas as pd
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

load_dotenv()

client = Groq(api_key=os.environ.get("GROQ_API_KEY"))

conn = sqlite3.connect("math_exam.py")
cursor = conn.cursor()

cursor.execute("""
               CREATE TABLE IF NOT EXISTS results (
                   id INTEGER PRIMARY KEY,
                   name TEXT,
                   grade TEXT,
                   difficulty TEXT,
                   score REAL,
                   date TEXT,
                   time_seconds REAL
               )
"""
)


name = input("What's your name? ")
grade = input("What's your grade level? ")


difficulty = input("Please choose your difficulty: easy, medium or hard: ")
while difficulty not in ["easy", "medium", "hard"]:
    difficulty = input("Please type easy, medium or hard: ")



response = client.chat.completions.create(
    model="llama-3.3-70b-versatile",
    messages=[{
        "role": "user", 
        "content": f"Generate 10 math questions for a grade/level {grade} student, use a {difficulty} difficulty level . Return ONLY a JSON array like this: [{{'question': 'What is 2+2?', 'answer': '4'}}]. No extra text, just the JSON."
    }]
)

questions = ast.literal_eval(response.choices[0].message.content)
correct = 0
total_time = 0
results = []
for q in questions:
    while True:
        start = time.time()
        answer = input(q["question"] + " ")
        end = time.time()
        elapsed = end - start
        total_time += elapsed
        if not answer:
            print("You have to type something")
        else:
            break
        
    results.append({
        "question": q["question"],
        "student_answer": answer,
        "correct_answer": q["answer"],
        "is_correct": answer == q["answer"],
        "time_seconds": round(elapsed, 2)
    })
    if answer == q["answer"]:
        print("Correct")
        correct += 1
    else:
        print("Incorrect")


        
score = correct / len(questions)
print(f"Your score is {score:.0%}")

df_results = pd.DataFrame(results)

summary = f"""

Here are your results {name}:
{df_results}
"""
print(summary)

feedback = client.chat.completions.create(
    model="llama-3.3-70b-versatile",
    messages=[{
        "role": "user",
        "content": f"A student named {name} in grade {grade} just took a math quiz. Using these results: {summary}, tell the student which topics they need to study and give them a personalized study guide about the topics they failed."
    }]
)

print(feedback.choices[0].message.content)
filename=f"{name}_results.csv"
df_results.to_csv(filename, index=False) 


#EXPORT RESULTS ASS EXCEL FILE

def export_to_excel(name, grade, difficulty, score, df_results, feedback_text, filename):
    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws1 = wb.active
    ws1.title = "Summary"
    ws1["A1"] = "Name"
    ws1["B1"] = name
    ws1["A2"] = "Grade"
    ws1["B2"] = grade
    ws1["A3"] = "Difficulty"
    ws1["B3"] = difficulty
    ws1["A4"] = "Score"
    ws1["B4"] = f"{score:.0%}"
    ws1["A5"] = "Date"
    ws1["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    for row in range(1, 6):
        ws1[f"A{row}"].font = Font(bold=True)

    ws1["A7"] = "Feedback"
    ws1["A7"].font = Font(bold=True)
    ws1["A8"] = feedback_text
    ws1["A8"].alignment = Alignment(wrap_text=True, vertical="top")
    ws1.merge_cells("A8:E20")
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 20

    # --- Sheet 2: Results ---
    ws2 = wb.create_sheet("Results")
    headers = ["Question", "Student Answer", "Correct Answer", "Correct?"]
    colors = ["FF0000", "00FF00", "0000FF", "FFA500"]

    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=colors[col - 1], end_color=colors[col - 1], fill_type="solid")

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for row_idx, r in enumerate(df_results.itertuples(index=False), start=2):
        ws2.cell(row=row_idx, column=1, value=r.question)
        ws2.cell(row=row_idx, column=2, value=r.student_answer)
        ws2.cell(row=row_idx, column=3, value=r.correct_answer)
        result_cell = ws2.cell(row=row_idx, column=4, value="Yes" if r.is_correct else "No")
        result_cell.fill = green if r.is_correct else red

    widths = [50, 20, 20, 12]
    for i, w in enumerate(widths, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(filename)
    print(f"\nResults saved to {filename}")

export_to_excel(name, grade, difficulty, score, df_results, feedback.choices[0].message.content, f"{name}_results.xlsx")